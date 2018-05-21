####ile Name: Ada_Code_File_Creation.py
#Author   : Rubber-Duck-999
#Version  : 1.0
#Status   : In Production

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.compat import range
import tkinter as tk
from openpyxl import load_workbook
from shutil import copyfile
import fileinput
#Import of necessary libraries

Main_2 = ("C:/Users/simon/Documents/Programming/Ada/System Checks")
Main   = ("C:/Users/simon/Documents/Github/Projects/ADLAS_Ada_System")  

Template = (Main + '/' + "System Project/Template Header.adb")

Project_Software_List = [(Main + '/' + "Software-Requirements/BIT"),
                         (Main + '/' + "Software-Requirements/ADLAS.Initialise_Target"),
                         (Main + '/' + "Software-Requirements/ADLAS.Phase_One_Analyse")]

Software_Requirements_List = [(Main + '/' + "System Project/BIT"),
                              (Main + '/' + "System Project/ADLAS.Initialise_Target"),
                              (Main + '/' + "System Project/ADLAS.Phase_One_Analyse")]

Project_Names = ["BIT",
                 "ADLAS.Initialise_Target",
                 "ADLAS.Phase_One_Analyse"]


def Project_Selection():
    print("#--------------------------------------------------------#")
    print("   Project Selection for ada autogeneration tool         ")
    print("#--------------------------------------------------------#")
    Index = 1
    for Project in Project_Names:
        print(str(Index) + ": Project Name: ", Project)
        Index += 1
    Index =- 1
    Choice = int(input("Select a project number?: "))
    print("#--------------------------------------------------------#")
    print("Project Name: " + Project_Names[Choice])
    print("#--------------------------------------------------------#")
    return Choice
    

def Replace(Ada, Function_Name):
    with open(Ada, "r") as File:
        print(Ada)
        Replacements_1 = File.read().replace('<File_Name>', Ada)
        Replacements_2 = File.read().replace('<Function_Name>', Function_Name)
    with open(Ada, "w") as File:
        File.write(Replacements_1)
        File.write(Replacements_2)


def Package_Body_A_Spec(File_Name_Lower_1, File_Name_Lower_2):
    Package_Body = (File_Name_Lower_1 + '-' + File_Name_Lower_2 + '.adb')
    Package_Spec = (File_Name_Lower_1 + '-' + File_Name_Lower_2 + '.ads')
    none = 'N/A'
    if (os.path.isfile(Package_Body)) == False:
        File = open((Package_Body), "w")
        with open(Template, 'r') as f:
            for line in f:
                File.write(line)
        File.close()
        print('Creating Package Body')
    Replace(Package_Body, none)    
    if (os.path.isfile(Package_Spec)) == False:
        File = open((Package_Spec), "w")
        with open(Template, 'r') as f:
            for line in f:
                File.write(line)
        File.close()
        Replace(Package_Spec, none)
        print('Creating Package Spec')

def Workbook(Choice):
    Req_Path = Software_Requirements_List[Choice]
    Proj_Path = Project_Software_List[Choice]
    os.chdir(Req_Path)
    Workbook_Name = input("Workbook Name? ")
    Workbook = load_workbook(Workbook_Name + '.xlsx')
    # grab the active worksheet
    Worksheet = Workbook.active
    os.chdir(Proj_Path)
    Column_Group = int(input("Which column for Group?: "))
    Column_Sub_Group = int(input("Which column for Sub_Group?: "))
    Column_Procedure = int(input("Which column for Procedures?: "))
    # Data can be assigned directly to cells
    Row = 2    
    while Worksheet.cell(row=Row, column=Column_Procedure).value is not None:
        os.chdir(Proj_Path)
        File_Name_Procedure = Worksheet.cell(row=Row, column=Column_Procedure).value
        File_Name_Sub_Group = Worksheet.cell(row=Row, column=Column_Sub_Group).value
        File_Name_Group = Worksheet.cell(row=Row, column=Column_Group).value
        File_Name_Lower_1 = File_Name_Group.lower()
        if (os.path.isdir('./' + File_Name_Lower_1)) == True:
            print('It Exists')
            Path = (Main_Path + '/' + File_Name_Lower_1 + '/')
            os.chdir(Path)
        else:
            print('Creating Group Folder')
            Make = ('./' + File_Name_Lower_1)
            os.mkdir(Make)
            Path = (Main_Path + '/' + File_Name_Lower_1)
            os.chdir(Path)
            
        File_Name_Lower_2 = File_Name_Sub_Group.lower()
        if (os.path.isdir('./' + File_Name_Lower_2)) == True:
            print('It Exists')
            Path = (Main_Path + '/' + File_Name_Lower_1 + '/' + File_Name_Lower_2)
            os.chdir(Path)
        else:
            print('Creating Sub Group Folder')
            Make = ('./' + File_Name_Lower_2)
            os.mkdir(Make)
            Path = (Main_Path + '/' + File_Name_Lower_1 + '/' + File_Name_Lower_2)
            os.chdir(Path)   
        Package_Body_A_Spec(File_Name_Lower_1, File_Name_Lower_2)    
        File_Name_Lower_3 = File_Name_Procedure.lower()
        if (os.path.isfile('./' + File_Name_Lower_3)) == True:
            print('It Exists')
        else:
            print('Creating File')
            Path = (Main_Path + '/' + File_Name_Lower_1 + '/' + File_Name_Lower_2)
            os.chdir(Path)
            File = open((File_Name_Lower_1 + "-" + File_Name_Lower_2 + "-" + File_Name_Lower_3 + ".adb"), "w")
            with open(Template, 'r') as f:
                for line in f:
                    File.write(line)
                #Replace(File)
        Row = Row + 1
        File.close()
        
    # Save the file
    Workbook.save(Workbook_Name + 'xlsx')

Choice = Project_Selection()
Workbook(Choice)



        
